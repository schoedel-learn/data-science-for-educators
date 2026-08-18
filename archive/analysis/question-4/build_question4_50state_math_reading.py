#!/usr/bin/env python3
from __future__ import annotations

import re
import subprocess
import urllib.request
from io import BytesIO
from pathlib import Path

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook

BASE = Path('/home/barry-schoedel/Documents/Obsidian Vault/🎓 Education/STEM-5328/Question 4 Exploration')
BASE.mkdir(parents=True, exist_ok=True)

CSV_PATH = BASE / 'question-4-charter-share-vs-naep-2024-50states.csv'
PNG_PATH = BASE / 'question-4-charter-share-vs-naep-2024-50states.png'
MD_PATH = BASE / 'question-4-charter-share-vs-naep-2024-50states.md'

CHARTER_XLSX = 'https://nces.ed.gov/programs/digest/d23/tables/xls/tabn216.90.xlsx'
STATE_CODES = 'AL,AK,AZ,AR,CA,CO,CT,DE,FL,GA,HI,ID,IL,IN,IA,KS,KY,LA,ME,MD,MA,MI,MN,MS,MO,MT,NE,NV,NH,NJ,NM,NY,NC,ND,OH,OK,OR,PA,RI,SC,SD,TN,TX,UT,VT,VA,WA,WV,WI,WY'
ALL_50_STATES = [
    'Alabama','Alaska','Arizona','Arkansas','California','Colorado','Connecticut','Delaware','Florida','Georgia',
    'Hawaii','Idaho','Illinois','Indiana','Iowa','Kansas','Kentucky','Louisiana','Maine','Maryland','Massachusetts',
    'Michigan','Minnesota','Mississippi','Missouri','Montana','Nebraska','Nevada','New Hampshire','New Jersey',
    'New Mexico','New York','North Carolina','North Dakota','Ohio','Oklahoma','Oregon','Pennsylvania','Rhode Island',
    'South Carolina','South Dakota','Tennessee','Texas','Utah','Vermont','Virginia','Washington','West Virginia',
    'Wisconsin','Wyoming'
]


def fetch_charter_share_all_50() -> pd.DataFrame:
    with urllib.request.urlopen(CHARTER_XLSX) as resp:
        data = resp.read()
    wb = load_workbook(filename=BytesIO(data), data_only=True)
    ws = wb.active

    charter = {}
    for row in ws.iter_rows(min_row=5, max_row=56, values_only=True):
        state = str(row[0]).replace('\xa0', ' ').strip() if row[0] is not None else ''
        if state in {'United States', 'District of Columbia', ''}:
            continue
        raw = row[16]  # 2022-23 charter enrollment share
        if raw == '#':
            charter[state] = 0.0
        elif raw in ('†', '---', None):
            # For the assignment we include all 50 states; states with no charter legislation / no charters
            # are coded as 0% charter enrollment so every state can appear as one dot.
            charter[state] = 0.0
        else:
            charter[state] = float(raw)

    # guarantee all 50 states present
    for state in ALL_50_STATES:
        charter.setdefault(state, 0.0)

    return pd.DataFrame({'state': ALL_50_STATES, 'charter_enroll_pct_2022_23': [charter[s] for s in ALL_50_STATES]})


def fetch_naep(subject: str) -> pd.DataFrame:
    cmd = [
        '/home/barry-schoedel/bin/naep-search',
        'data',
        subject,
        '8',
        '-y',
        '2024',
        '-g',
        'TOTAL',
        '-t',
        'MN:MN',
        '-j',
        STATE_CODES,
    ]
    out = subprocess.check_output(cmd, text=True)
    pat = re.compile(r'^(\d{4})\s+([A-Z]{2})\s+(.+?)\s+TOTAL\s+All students\s+1\s+All students\s+([0-9.]+)\s+1\s+0\s*$')
    rows = []
    for line in out.splitlines()[2:]:
        m = pat.match(line)
        if m:
            _, code, state, value = m.groups()
            rows.append({'state': state.strip(), f'naep_{subject}_2024': float(value), 'state_code': code})
    return pd.DataFrame(rows)


def summarize(ax, df, ycol: str, title: str) -> tuple[float, float]:
    x = df['charter_enroll_pct_2022_23'].to_numpy()
    y = df[ycol].to_numpy()
    corr = float(np.corrcoef(x, y)[0, 1])
    slope, intercept = np.polyfit(x, y, 1)
    xs = np.linspace(x.min(), x.max(), 200)

    zero_mask = df['charter_enroll_pct_2022_23'] == 0.0
    ax.scatter(x[~zero_mask], y[~zero_mask], s=42, alpha=0.85, color='#2563eb', edgecolor='white', linewidth=0.5, label='States with charter share > 0%')
    ax.scatter(x[zero_mask], y[zero_mask], s=52, alpha=0.95, color='#f59e0b', marker='s', edgecolor='white', linewidth=0.6, label='States coded at 0% charter share')
    ax.plot(xs, slope * xs + intercept, color='#dc2626', linewidth=2, label='Best-fit line')
    ax.set_title(f'{title}\nr = {corr:.2f}, slope = {slope:.2f}\nRaw, unadjusted state-level relationship; exploratory only', fontsize=11)
    ax.set_xlabel('Charter enrollment share (2022–23)')
    ax.set_ylabel('NAEP grade 8 mean score (2024)')
    ax.grid(True, linestyle='--', alpha=0.25)
    ax.legend(frameon=False, fontsize=8, loc='best')

    for state in ['Arizona', 'Massachusetts', 'Texas', 'Colorado', 'New Mexico', 'Florida']:
        row = df[df['state'] == state]
        if not row.empty:
            ax.annotate(state, (row.iloc[0]['charter_enroll_pct_2022_23'], row.iloc[0][ycol]), xytext=(4, 4), textcoords='offset points', fontsize=8)

    return corr, slope


def main() -> None:
    charter = fetch_charter_share_all_50()
    math_df = fetch_naep('math').drop(columns=['state_code'])
    reading_df = fetch_naep('reading').drop(columns=['state_code'])

    df = charter.merge(math_df, on='state', how='left').merge(reading_df, on='state', how='left')
    df = df.sort_values('charter_enroll_pct_2022_23').reset_index(drop=True)
    df.to_csv(CSV_PATH, index=False)

    fig, axes = plt.subplots(1, 2, figsize=(14, 6), constrained_layout=True)
    math_corr, math_slope = summarize(axes[0], df, 'naep_math_2024', 'Charter Share vs. NAEP Math')
    read_corr, read_slope = summarize(axes[1], df, 'naep_reading_2024', 'Charter Share vs. NAEP Reading')
    fig.suptitle('Question 4 Exploratory Scatterplots — All 50 States', fontsize=15)
    fig.savefig(PNG_PATH, dpi=220)
    plt.close(fig)

    no_charter_states = [s for s in ALL_50_STATES if float(df.loc[df['state'] == s, 'charter_enroll_pct_2022_23'].iloc[0]) == 0.0]
    nonzero = df[df['charter_enroll_pct_2022_23'] > 0].copy()
    math_corr_nz = float(nonzero['charter_enroll_pct_2022_23'].corr(nonzero['naep_math_2024']))
    read_corr_nz = float(nonzero['charter_enroll_pct_2022_23'].corr(nonzero['naep_reading_2024']))

    MD_PATH.write_text(f'''---
type: analysis-note
title: "Question 4 — 50-State 2024 Charter Share vs. NAEP Scatterplots"
created: 2026-08-02
question: "What is the relationship between state-level charter school penetration and aggregate NAEP performance, controlling for demographic composition?"
source_1: "NAEP Data Service API (2024 grade 8 math and reading state means)"
source_2: "NCES Digest table 216.90 (2022–23 charter enrollment share by state)"
---

# Question 4 — 50-State 2024 Scatterplots

Files created:
- CSV: `{CSV_PATH.name}`
- Chart: `{PNG_PATH.name}`

## Accuracy correction from the earlier draft

The earlier exploratory pass used 44 states because five no-charter states were omitted as “not applicable,” and Wyoming was accidentally dropped by an Excel row-slice bug. This corrected version includes **all 50 states**.

To make that possible, states reported by NCES as having **no charter share in 2022–23** were coded as **0% charter enrollment**. This includes both states with no charter legislation / no operating charter schools in the table and Kentucky, which is listed at 0.0%.

States at 0% charter enrollment in this file:
- {', '.join(no_charter_states)}

## Math (2024)
- Correlation: **{math_corr:.2f}**
- Slope: **{math_slope:.2f}** NAEP math points per 1 percentage-point increase in charter share

Plain-language reading:
- The raw relationship is **negative**, but weak.
- Charter share alone shows only a **weak linear relationship** with state math performance.

Sensitivity check:
- Excluding the 0% charter-share states, the relationship stays negative but becomes smaller (**r = {math_corr_nz:.2f}**).

## Reading (2024)
- Correlation: **{read_corr:.2f}**
- Slope: **{read_slope:.2f}** NAEP reading points per 1 percentage-point increase in charter share

Plain-language reading:
- The raw relationship is **negative**, but weak.
- Charter share alone shows only a **weak linear relationship** with state reading performance.

Sensitivity check:
- Excluding the 0% charter-share states, the relationship stays negative but becomes smaller (**r = {read_corr_nz:.2f}**).

## Interpretation

These are exploratory, raw scatterplots. They show whether there is a visible relationship worth investigating, not whether charter schools cause better or worse scores. The next step is to add controls for poverty, race/ethnicity, disability, English-learner status, and possibly spending.
''')

    print(f'CSV: {CSV_PATH}')
    print(f'PNG: {PNG_PATH}')
    print(f'MD:  {MD_PATH}')
    print(f'Math correlation: {math_corr:.4f}')
    print(f'Math slope: {math_slope:.4f}')
    print(f'Reading correlation: {read_corr:.4f}')
    print(f'Reading slope: {read_slope:.4f}')


if __name__ == '__main__':
    main()
